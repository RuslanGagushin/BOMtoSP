import openpyxl

# Загрузка исходного BOM листа
wb = openpyxl.load_workbook(
    'Z:/Проекты/Текущие/395-396 Вирион/2. Проектирование/2.2 BOM/BTCH.395.652.21.BOM Реактор R650L Rev.4.xlsx')

ws = wb.active  # рабочий лист в книге с BOM листом

# Создание спецификации
nb = openpyxl.Workbook()
file_name = "Specification.xlsx"  # Filename of SP
ns = nb.active  # New sheet for SP

# Сбор Наименований
for i in range(10, 300):
    if ws[f"B{i + 1}"].value is not None:
        ns[f"A{i - 9}"] = ws[f"B{i + 1}"].value
    else:
        break

# Сбор Производителей
for i in range(10, 300):
    if ws[f"D{i + 1}"].value is not None:
        ns[f"C{i - 9}"] = ws[f"D{i + 1}"].value
    else:
        break

# Сбор кода продукта
for i in range(10, 300):
    if ws[f"E{i + 1}"].value is not None:
        ns[f"B{i - 9}"] = ws[f"E{i + 1}"].value
    else:
        break

# Добавление количества
for i in range(10, 300):
    if ws[f"E{i + 1}"].value == "Код продукта":
        ns[f"D{i - 9}"] = "Кол-во"
    elif ws[f"E{i + 1}"].value is not None:
        ns[f"D{i - 9}"] = 1
    else:
        break

# Объединение Наименования и кода продукта
for i in range(0, 300):
    if ns[f"A{i + 1}"].value is not None:
        ns[f"A{i + 1}"] = ns[f"A{i + 1}"].value + " " + ns[f"B{i + 1}"].value
    else:
        break

ns.delete_cols(2)


nb.save('Z:/Проекты/Текущие/395-396 Вирион/2. Проектирование/2.2 BOM/Spec.xlsx')